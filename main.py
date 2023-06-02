import os
import pathlib
from tkinter.ttk import Labelframe

import win32com.client
from os import path
from openpyxl import load_workbook
from openpyxl import Workbook
import tkinter as tk
import tkinter.filedialog
import ttkbootstrap as ttk
from ttkbootstrap.scrolled import ScrolledText
from ttkbootstrap.constants import *
from tkinter.filedialog import askdirectory, askopenfilename
from ttkbootstrap.dialogs import Messagebox, Querybox
from pathlib import Path
import datetime
import configparser

red_sentences = []


class WordMaker(ttk.Frame):
    def __init__(self, master, **Kwargs):
        super().__init__(master, **Kwargs)
        self.pack(fill=BOTH, expand=YES)
        self.wordTemplate_path = ttk.StringVar()
        self.excel_path = ttk.StringVar()
        self.folder_path = ttk.StringVar()
        self.prog_tab1 = ttk.IntVar()
        self.prog_tab2 = ttk.IntVar()
        self.prog_tab3 = ttk.IntVar()
        self.rowRange1 = ttk.IntVar()
        self.rowRange2 = ttk.IntVar()
        self.folder_path2 = ttk.StringVar()
        self.folder_path3 = ttk.StringVar()
        self.wdColor = ttk.IntVar()  # 模板标记颜色
        self.appConfig = ttk.StringVar()  # office/wps
        self.config = configparser.ConfigParser()
        self.appConfig_dic = {"kWPS": "kwps.Application", "WPS": "wps.Application", "Office": "Word.Application"}
        self.color_dic = {"红色": 255, "自动(黑)": -16777216, "黑色": 0, "蓝色": 16711680, "金色": 52479,
                     "绿色": 32768, "黄色": 65535}

        self.loadConfig()

        # Frame configue
        # self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=1)
        # self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        hdr_frame = ttk.Frame(self, padding=20, bootstyle=SECONDARY)
        hdr_frame.grid(row=0, column=0, columnspan=2, sticky=NSEW)
        # hdr_label = ttk.Label(
        #     master=hdr_frame,
        #     image='logo',
        #     bootstyle=(INVERSE, SECONDARY)
        # )
        # hdr_label.pack(side=LEFT, expand=YES)

        logo_text = ttk.Label(
            master=hdr_frame,
            text='Word Maker',
            font=('TkDefaultFixed', 30),
            bootstyle=(INVERSE, SECONDARY)
        )
        logo_text.pack(side=LEFT, padx=10, expand=YES)

        # option notebook
        self.nb = ttk.Notebook(self, bootstyle=LIGHT)
        self.nb.grid(row=1, column=1, sticky=NSEW)
        nb1 = ttk.Frame(self.nb)
        self.nb.add(
            # child=ttk.Label(nb, text="tab content"),
            child=nb1,
            # text="Tab 1",
            sticky=NW
        )
        nb2 = ttk.Frame(self.nb)
        self.nb.add(
            child=nb2,
            # text="Tab 2",
            sticky=NW
        )
        nb3 = ttk.Frame(self.nb)
        self.nb.add(
            child=nb3,
            # text="Tab 3",
            sticky=NW
        )
        nb4 = ttk.Frame(self.nb)
        self.nb.add(
            child=nb4,
            # text="Tab 4",
            sticky=NW
        )
        nb5 = ttk.Frame(self.nb)
        self.nb.add(
            child=nb5,
            # text="Tab 5",
            sticky=NW
        )

        # notebook tab1
        self.notebook1(nb1)
        self.notebook2(nb2)
        self.notebook3(nb3)
        self.notebook4(nb4)
        self.notebook5(nb5)

        # action buttons
        action_frame = ttk.Frame(self)
        action_frame.grid(row=1, column=0, sticky=NSEW)

        tab1_btn = ttk.Button(
            master=action_frame,
            # image='cleaner',
            text='excel builder',
            compound=TOP,
            command=lambda: self.nb.select(0),
            bootstyle=INFO
        )
        tab1_btn.pack(side=TOP, fill=BOTH, ipadx=10, ipady=10, expand=YES)

        tab2_btn = ttk.Button(
            master=action_frame,
            # image='registry',
            text='excel to word',
            compound=TOP,
            command=lambda: self.nb.select(1),
            bootstyle=INFO
        )
        tab2_btn.pack(side=TOP, fill=BOTH, ipadx=10, ipady=10, expand=YES)

        tab3_btn = ttk.Button(
            master=action_frame,
            # image='tools',
            text='word to PDF',
            compound=TOP,
            command=lambda: self.nb.select(2),
            bootstyle=INFO
        )
        tab3_btn.pack(side=TOP, fill=BOTH, ipadx=10, ipady=10, expand=YES)

        tab4_btn = ttk.Button(
            master=action_frame,
            # image='options',
            text='about',
            compound=TOP,
            command=lambda: self.nb.select(3),
            bootstyle=INFO
        )
        tab4_btn.pack(side=TOP, fill=BOTH, ipadx=10, ipady=10, expand=YES)

        tab5_btn = ttk.Button(
            master=action_frame,
            # image='options',
            text='options',
            compound=TOP,
            command=lambda: self.nb.select(4),
            bootstyle=INFO
        )
        tab5_btn.pack(side=TOP, fill=BOTH, ipadx=10, ipady=10, expand=YES)





    def notebook1(self, parent):
        container = ttk.Frame(parent)
        container.grid(row=0, column=0, columnspan=3, sticky=NSEW, padx=80, pady=20)
        # word file input
        word_label = ttk.Label(
            master=container,
            font="TkFixedFont 10",
            text="导入Word标记模板(.docx)",
            anchor=NW,
        )
        word_label.grid(row=1, column=0, columnspan=3, sticky=NSEW, pady=10)

        file_entry = ttk.Entry(container, textvariable=self.wordTemplate_path)
        file_entry.grid(row=2, column=0, columnspan=2, sticky=NSEW)

        open_btn = ttk.Button(
            master=container,
            # image=
            text="open",
            bootstyle=(OUTLINE, SECONDARY),
            command=self.getWordPath
        )
        open_btn.grid(row=2, column=2, sticky=NSEW)
        # word file input
        excel_label = ttk.Label(
            master=container,
            font="TkFixedFont 10",
            text="选择导出的路径",
            anchor=NW,
        )
        excel_label.grid(row=3, column=0, columnspan=3, sticky=NSEW, pady=10)

        file_entry = ttk.Entry(container, textvariable=self.folder_path)
        file_entry.grid(row=4, column=0, columnspan=2, sticky=NSEW)

        open_btn = ttk.Button(
            master=container,
            # image=
            text="open",
            bootstyle=(OUTLINE, SECONDARY),
            command=self.getDirectory
        )
        open_btn.grid(row=4, column=2, sticky=NSEW)

        ok_btn = ttk.Button(
            master=container,
            text="制作Excel表单",
            compound=BOTTOM,
            bootstyle=SUCCESS,
            command=self.excelBuilder
        )
        ok_btn.grid(row=6, column=1, sticky=S, pady=10)

        pb = ttk.Progressbar(
            master=container,
            variable=self.prog_tab1,
            bootstyle=(SUCCESS, STRIPED),
        )
        pb.grid(row=7, column=0, columnspan=4, sticky=EW)
        self.prog_tab1.set(0)

    def notebook2(self, parent):
        container = ttk.Frame(parent)
        container.grid(row=0, column=0, columnspan=3, sticky=NSEW, padx=20, pady=5)
        # word file input
        word_label = ttk.Label(
            master=container,
            font="TkFixedFont 10",
            text="导入Word标红模板(.docx)",
            anchor=NW,
        )
        word_label.grid(row=1, column=0, columnspan=3, sticky=NSEW, pady=5)

        file_entry = ttk.Entry(container, textvariable=self.wordTemplate_path)
        file_entry.grid(row=2, column=0, columnspan=2, sticky=NSEW)

        open_btn = ttk.Button(
            master=container,
            # image=
            text="open",
            bootstyle=(OUTLINE, SECONDARY),
            command=self.getWordPath
        )
        open_btn.grid(row=2, column=2, sticky=NSEW)
        # word file input
        excel_label = ttk.Label(
            master=container,
            font="TkFixedFont 10",
            text="选择数据源(.xlsx)",
            anchor=NW,
        )
        excel_label.grid(row=3, column=0, columnspan=3, sticky=NSEW, pady=5)

        file_entry = ttk.Entry(container, textvariable=self.excel_path)
        file_entry.grid(row=4, column=0, columnspan=2, sticky=NSEW)

        open_btn = ttk.Button(
            master=container,
            # image=
            text="open",
            bootstyle=(OUTLINE, SECONDARY),
            command=self.getExcelPath
        )
        open_btn.grid(row=4, column=2, sticky=NSEW)

        input_group = ttk.Labelframe(container, text="range pick", bootstyle="info", padding=5)
        input_group.grid(row=5, column=0, columnspan=3, sticky=EW, ipadx=5)
        self.rowRange1.set(2)
        self.rowRange2.set(0)
        entry1 = ttk.Entry(input_group, textvariable=self.rowRange1)
        entry1.pack(side=LEFT)
        input_label = ttk.Label(input_group, text="~")
        input_label.pack(side=LEFT)
        entry2 = ttk.Entry(input_group, textvariable=self.rowRange2)
        entry2.pack(side=RIGHT)

        ok_btn = ttk.Button(
            master=container,
            text="开始执行",
            compound=BOTTOM,
            bootstyle=SUCCESS,
            command=self.excel2word_mul
        )
        ok_btn.grid(row=6, column=1, sticky=S, pady=2)

        pb = ttk.Progressbar(
            master=container,
            variable=self.prog_tab2,
            bootstyle=(SUCCESS, STRIPED),
        )
        pb.grid(row=7, column=0, columnspan=4, sticky=EW)
        self.prog_tab2.set(0)

    def notebook3(self, parent):
        container = ttk.Frame(parent)
        container.grid(row=0, column=0, columnspan=3, sticky=NSEW, padx=80, pady=20)
        # word file input
        word_label = ttk.Label(
            master=container,
            font="TkFixedFont 10",
            text="word文档存放路径(批量处理)",
            anchor=NW,
        )
        word_label.grid(row=1, column=0, columnspan=3, sticky=NSEW, pady=10)

        file_entry = ttk.Entry(container, textvariable=self.folder_path2)
        file_entry.grid(row=2, column=0, columnspan=2, sticky=NSEW)

        open_btn = ttk.Button(
            master=container,
            # image=
            text="open",
            bootstyle=(OUTLINE, SECONDARY),
            command=self.getDirectory2
        )
        open_btn.grid(row=2, column=2, sticky=NSEW)

        # word file input
        excel_label = ttk.Label(
            master=container,
            font="TkFixedFont 10",
            text="选择导出PDF存放路径",
            anchor=NW,
        )
        excel_label.grid(row=3, column=0, columnspan=3, sticky=NSEW, pady=10)

        file_entry = ttk.Entry(container, textvariable=self.folder_path3)
        file_entry.grid(row=4, column=0, columnspan=2, sticky=NSEW)

        open_btn = ttk.Button(
            master=container,
            # image=
            text="open",
            bootstyle=(OUTLINE, SECONDARY),
            command=self.getDirectory3
        )
        open_btn.grid(row=4, column=2, sticky=NSEW)

        ok_btn = ttk.Button(
            master=container,
            text="开始执行",
            compound=BOTTOM,
            bootstyle=SUCCESS,
            command=self.word2pdf
        )
        ok_btn.grid(row=5, column=1, sticky=S, pady=10)

        pb = ttk.Progressbar(
            master=container,
            variable=self.prog_tab3,
            bootstyle=(SUCCESS, STRIPED),
        )
        pb.grid(row=6, column=0, columnspan=4, sticky=EW)
        self.prog_tab3.set(0)

    def notebook4(self, parent):
        container = ttk.Frame(parent, padding=5)
        container.grid(row=0, column=0, sticky=NSEW)
        style = ttk.Style()
        textbox = ScrolledText(
            master=container,
            highlightcolor=style.colors.primary,
            highlightbackground=style.colors.border,
            highlightthickness=1,
            autohide=True,
            width=48,
            height=11,
        )
        textbox.grid(row=0, column=0, sticky=NS)
        default_txt = "帮助：\n" \
                      "1) 先确认自己word程序是WPS还是Office,并在options界面进行相应版本切换\n" \
                      "2) 在word里面给关键字字体进行颜色标记(默认红色)\n" \
                      "3) 在excel builder页面导入之前标记好的文档,然后制作excel数据源\n" \
                      "4) 在excel to word页面进行批量生成word,分别要导入模板word和数据源," \
                      "   选择excel行号区间设置生成数据的范围。\n" \
                      "5) 在word to PDF进行批量word转换PDF操作,只须选择相应的文件目录就可以了\n" \
                      "6) 注意: 执行过程前请先关闭相应的word/excel文件,执行过程中程序无响应是正常的," \
                      "   请耐心等待程序执行完成直到弹出提示框,默认支持的文件类型:.docx/.xlsx\n" \
                      "7) 关于配置文件：在设置的保存配置选项可以将当前选取的路径和设置选项保存在当前程序的本地目录下\n" \
                      "作者QQ:2496729202\n" \
                      "-------------------------------\n" \
                      "v1.1更新说明            2023/3/30\n" \
                      "- 适配WPS,增加版本切换设置\n" \
                      "v1.2更新说明            2023/4/4\n" \
                      "- 新增配置文件"
        textbox.insert(END, default_txt)
        textbox.text.configure(state="disabled")

    def notebook5(self, parent):
        container = ttk.Frame(parent, padding=5)
        container.grid(row=0, column=0, sticky=NSEW)
        text_Label = ttk.Label(
            master=container,
            font="TkFixedFont 10",
            text="Settings:",
            anchor=NW,
        )
        text_Label.grid(row=1, column=0, sticky=NW)

        menu_markPick = ttk.Menu(container)

        for k in self.color_dic:
            menu_markPick.add_radiobutton(label=k, value=self.color_dic[k], variable=self.wdColor)

        mb = ttk.Menubutton(
            master=container,
            text="切换模板标记字体颜色",
            bootstyle=(WARNING, OUTLINE),
            menu=menu_markPick
        )
        mb.grid(row=2, column=0, sticky=NSEW)

        menu_appPick = ttk.Menu(container)

        for k in self.appConfig_dic:
            menu_appPick.add_radiobutton(label=k, value=self.appConfig_dic[k], variable=self.appConfig)

        mb2 = ttk.Menubutton(
            master=container,
            text="切换文档应用程序接口",
            bootstyle=(DANGER, OUTLINE),
            menu=menu_appPick
        )
        mb2.grid(row=3, column=0, sticky=NSEW, pady=20)

        btn_saveConfig = ttk.Button(
            master=container,
            text="保存配置状态",
            compound=BOTTOM,
            bootstyle=SUCCESS,
            command=self.saveConfig
        )
        btn_saveConfig.grid(row=4, column=0, sticky=NSEW, pady=20)

    def loadConfig(self):
        try:
            if pathlib.Path('config.ini').exists():
                self.config.read('config.ini',encoding='UTF-8')
                if "PathConfiguration" in self.config.sections():
                    if "word_template_path" in self.config["PathConfiguration"]:
                        self.wordTemplate_path.set(self.config["PathConfiguration"]["word_template_path"])
                    if "excel_output_folder" in self.config["PathConfiguration"]:
                        self.folder_path.set(self.config["PathConfiguration"]["excel_output_folder"])
                    if "word_template_path" in self.config["PathConfiguration"]:
                        self.excel_path.set(self.config["PathConfiguration"]["word_template_path"])
                    if "word_output_folder" in self.config["PathConfiguration"]:
                        self.folder_path2.set(self.config["PathConfiguration"]["word_output_folder"])
                    if "pdf_output_folder" in self.config["PathConfiguration"]:
                        self.folder_path3.set(self.config["PathConfiguration"]["pdf_output_folder"])
                if "Options" in self.config.sections():
                    if "default_marking_color" in self.config["Options"]:
                        self.wdColor.set(self.color_dic[self.config["Options"]["default_marking_color"]])
                    if "default_API" in self.config["Options"]:
                        self.appConfig.set(self.appConfig_dic[self.config["Options"]["default_API"]])
            else:
                self.saveConfig()
        except Exception as e:
            Messagebox.show_error(message="[配置文件初始化失败]"+str(e))


    def saveConfig(self):
        if "PathConfiguration" not in self.config:
            self.config.add_section("PathConfiguration")
        if self.wordTemplate_path.get() != "":
            self.config["PathConfiguration"]["word_template_path"] = self.wordTemplate_path.get()
        if self.folder_path.get() != "":
            self.config["PathConfiguration"]["excel_output_folder"] = self.folder_path.get()
        if self.excel_path.get() != "":
            self.config["PathConfiguration"]["excel_input_path"] = self.excel_path.get()
        if self.folder_path2.get() != "":
            self.config["PathConfiguration"]["word_output_folder"] = self.folder_path2.get()
        if self.folder_path3.get() != "":
            self.config["PathConfiguration"]["pdf_output_folder"] = self.folder_path3.get()
        if "Options" not in self.config:
            self.config.add_section("Options")
        if self.wdColor.get() == 0:
            self.wdColor.set(255)
            for k in self.color_dic.keys():
                if self.color_dic[k] == self.wdColor.get():
                    self.config["Options"]["default_marking_color"] = k
                    break
        else:
            for k in self.color_dic.keys():
                if self.color_dic[k] == self.wdColor.get():
                    self.config["Options"]["default_marking_color"] = k
                    break
        if self.appConfig.get() == "":
            self.appConfig.set("kwps.Application")
            for k in self.appConfig_dic:
                if self.appConfig_dic[k] == self.appConfig.get():
                    self.config["Options"]["default_API"] = k
        else:
            for k in self.appConfig_dic:
                if self.appConfig_dic[k] == self.appConfig.get():
                    self.config["Options"]["default_API"] = k
        with open('config.ini', 'w', encoding='UTF-8') as configfile:
            self.config.write(configfile)

    def getWordPath(self):
        self.update_idletasks()
        d = askopenfilename()
        if d and (d.endswith(".docx") or d.endswith(".doc")):
            self.wordTemplate_path.set(d)  # self.setvar('word-path', d)
        else:
            Messagebox.show_error(message='请选择正确的文档格式!(.docx/.doc)')

    def getExcelPath(self):
        self.update_idletasks()
        d = askopenfilename()
        if d and d.endswith(".xlsx"):
            self.excel_path.set(d)  # self.setvar('word-path', d)
        else:
            Messagebox.show_error(message='请选择正确的文档格式!(.xlsx)')

    def getDirectory(self):
        self.update_idletasks()
        d = askdirectory()
        if d:
            self.folder_path.set(d)

    def getDirectory2(self):
        self.update_idletasks()
        d = askdirectory()
        if d:
            self.folder_path2.set(d)

    def getDirectory3(self):
        self.update_idletasks()
        d = askdirectory()
        if d:
            self.folder_path3.set(d)

    def excelBuilder(self):
        # Messagebox.show_info(message='注意:执行程序前请关闭目标文档，否则会导致进程占用!')
        self.prog_tab1.set(10)
        docxPath = self.wordTemplate_path.get()
        xlsxPath = self.folder_path.get()
        if docxPath == "" or xlsxPath == "":
            Messagebox.show_error(message='[错误]:请选择正确的路径!')
            self.prog_tab1.set(0)
            return
        xlsxPath = Path(xlsxPath) / (Path(docxPath).name.split(".")[0] + ".xlsx")
        try:
            word = win32com.client.DispatchEx(self.appConfig.get())
            word.Visible = 0  # 后台运行
            word.DisplayAlerts = 0  # 不显示，不警告
            doc = word.Documents.Open(docxPath)  # 打开一个已有的word文档
        except Exception as e:
            # Messagebox.show_error(message='[错误]:进程被占用,执行程序前请关闭目标文档!')
            Messagebox.show_error(message='[错误]:' + str(e))
            self.prog_tab1.set(0)
            return
        self.prog_tab1.set(30)
        find = doc.Range().Find
        find.ClearFormatting()
        find.Font.Color = self.wdColor.get()
        while find.Execute():
            red_text = find.Parent.Text
            if str(red_text).endswith("\r"):
                find.Parent.Text = red_text.replace("\r", "")
            red_sentences.append(red_text)
            find.Forward

        # print(red_sentences)
        doc.Close()
        self.prog_tab1.set(60)
        columns = []
        for i in range(len(red_sentences)):
            columns.append(f"col_{i}")

        wb = Workbook()
        ws = wb.active
        ws.append(columns)
        ws.append(red_sentences)
        wb.save(xlsxPath)
        self.prog_tab1.set(100)
        Messagebox.ok(message='已完成!')
        self.prog_tab1.set(0)

    def excel2word_mul(self):
        self.prog_tab2.set(0)
        row1 = self.rowRange1.get()
        row2 = self.rowRange2.get()
        path_xlsx = self.excel_path.get()
        docxPath = self.wordTemplate_path.get()
        if docxPath == "" or path_xlsx == "":
            Messagebox.show_error(message='[错误]:请选择正确的路径!')
            self.prog_tab2.set(0)
            return
        try:
            wb = load_workbook(path_xlsx)
            ws = wb.active
        except:
            Messagebox.show_error(message='[错误]:请选择正确的路径!')
            self.prog_tab2.set(0)
            return
        self.prog_tab2.set(10)
        if row2 < row1:
            Messagebox.show_error(message='[错误]:请输入正确的表单行号范围')
            self.prog_tab2.set(0)
            return

        self.prog_tab2.set(20)
        minRow = ws.min_row
        maxRow = ws.max_row
        if row1 < minRow or row2 > maxRow:
            Messagebox.show_error(message='[错误]:请输入正确的表单行号范围')
            return

        selected = [[] for i in range(row1, row2 + 1)]
        for i, ss in zip(range(row1, row2 + 1), selected):
            for cell in ws[i]:
                ss.append(cell.value)

        self.prog_tab2.set(30)
        prog_cnt = 30
        cnt = 0
        for ss in selected:
            try:
                word = win32com.client.DispatchEx(self.appConfig.get())
                word.Visible = 0  # 后台运行
                word.DisplayAlerts = 0  # 不显示，不警告
                doc = word.Documents.Open(docxPath)  # 打开一个已有的word文档
            except Exception as e:
                Messagebox.show_error(message='[错误]:' + str(e))
                return

            rg = doc.Range()
            rg.Find.ClearFormatting()
            rg.Find.Font.Color = self.wdColor.get()
            i = 0
            while rg.Find.Execute("", False, False, False, False, False, True, 1, True, ss[i], 1):
                rg.Font.Color = 0
                if i + 1 == len(ss):
                    break
                i += 1
                rg.Find.Forward

            now = datetime.datetime.now()
            timestamp = now.strftime("%Y-%m-%d")
            docxPath_sav = Path(
                docxPath).parent / f'output_{timestamp}-{str(Path(docxPath).name).split(".")[0]}' / f'{cnt}.docx'
            # docxPath_sav = Path(docxPath).parent / f'{cnt}.docx'
            if not os.path.exists(str(docxPath_sav.parent)):
                os.mkdir(str(docxPath_sav.parent))
            # print(str(docxPath_sav))
            doc.SaveAs(str(docxPath_sav))
            doc.Close()
            word.Quit()
            prog_cnt += int(cnt / len(selected) * 70)
            cnt += 1
            if prog_cnt <= 100:
                self.prog_tab2.set(prog_cnt)

        self.prog_tab2.set(100)
        Messagebox.ok(title="完成", message='已生成的文件夹在模板的同级目录')

    # def excel2word_single(self):
    #     path_xlsx = self.excel_path.get()
    #     try:
    #         wb = load_workbook(path_xlsx)
    #         ws = wb.active
    #     except:
    #         Messagebox.show_error(message='[错误]:请选择正确的路径!')
    #     rowNum = Querybox.get_integer(prompt="输入数据源行号", minvalue=2)
    #     selected = []
    #     for cell in ws[f"{rowNum}"]:
    #         selected.append(cell.value)
    #
    #     word = win32com.client.DispatchEx('Word.Application')
    #     word.Visible = 0  # 后台运行
    #     word.DisplayAlerts = 0  # 不显示，不警告
    #     doc = word.Documents.Open(path_doc)  # 打开一个已有的word文档
    #     rg = doc.Range()
    #     rg.Find.ClearFormatting()
    #     rg.Find.Font.Color = self.wdColor.get()
    #     i = 0
    #     while rg.Find.Execute("", False, False, False, False, False, True, 1, True, selected[i], 1):
    #         rg.Font.Color = 0
    #         if i + 1 == len(selected):
    #             break
    #         i += 1
    #         rg.Find.Forward
    #
    #     doc.SaveAs(path_doc1)
    #     doc.Close()

    def word2pdf(self):
        self.prog_tab3.set(0)
        wordPath = self.folder_path2.get()
        excelPath = self.folder_path3.get()
        if wordPath == "" or excelPath == "":
            Messagebox.show_error(message='[错误]:请选择正确的路径!')
            self.prog_tab3.set(0)
            return
        wordPath = Path(wordPath)
        out_files = []
        for file in wordPath.rglob("*.docx"):
            out_files.append(str(Path(excelPath) / file.name))
        if len(out_files) == 0:
            Messagebox.show_error(message='[错误]:该路径下没有.docx文档!')
            self.prog_tab3.set(0)
            return
        self.prog_tab3.set(10)
        cnt = 0
        prog_cnt = 10
        for f in out_files:
            try:
                word = win32com.client.DispatchEx(self.appConfig.get())
                word.Visible = 0  # 后台运行
                word.DisplayAlerts = 0  # 不显示，不警告
                doc = word.Documents.Open(f)  # 打开一个已有的word文档
                doc.SaveAs(f.replace(".docx", ".pdf"), FileFormat=17)
                doc.Close()
                word.Quit()
            except Exception as e:
                Messagebox.show_error(message='[错误]:' + str(e))
                return
            prog_cnt += int(cnt / len(out_files) * 90)
            if prog_cnt <= 100:
                self.prog_tab3.set(prog_cnt)

        self.prog_tab3.set(100)
        Messagebox.ok(title="完成", message='生成的pdf文件在word的同级目录下')


if __name__ == "__main__":
    app = ttk.Window(
        title="WordMaker V1.2",
        themename="flatly",
        size=(600, 380),
        resizable=(False, False),
        iconphoto=str(Path(__file__).parent / "icon.ico")
    )
    WordMaker(app)
    app.mainloop()

"""pyinstaller
pyinstaller -F -w -i D:\Project\JetBrains_WorkSpace\PycharmProjects\WordMaker\icon.ico D:\Project\JetBrains_WorkSpace\PycharmProjects\WordMaker\main.py
"""
